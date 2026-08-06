from __future__ import annotations

import csv
import hashlib
import io
import json
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from typing import Iterable, Mapping, Protocol

from openpyxl import load_workbook


class FindingSeverity(StrEnum):
    WARNING = "warning"
    ERROR = "error"


@dataclass(frozen=True, slots=True)
class FinanceRecord:
    company: str
    account: str
    period: str
    scenario: str
    value: Decimal
    currency: str
    dimensions: tuple[tuple[str, str], ...] = ()
    source_row: int | None = None

    def canonical_key(self) -> tuple[object, ...]:
        return (
            self.company,
            self.account,
            self.period,
            self.scenario,
            self.currency,
            self.dimensions,
        )


@dataclass(frozen=True, slots=True)
class DataQualityFinding:
    code: str
    severity: FindingSeverity
    message: str
    row_number: int | None = None


@dataclass(frozen=True, slots=True)
class DataQualityReport:
    row_count: int
    findings: tuple[DataQualityFinding, ...]

    @property
    def blocking(self) -> bool:
        return any(item.severity == FindingSeverity.ERROR for item in self.findings)

    @property
    def score(self) -> float:
        if self.row_count == 0:
            return 0.0
        penalty = sum(10 if item.severity == FindingSeverity.ERROR else 2 for item in self.findings)
        return max(0.0, round(100.0 - penalty / self.row_count, 2))


@dataclass(frozen=True, slots=True)
class DataSnapshot:
    snapshot_id: str
    content_hash: str
    row_count: int
    records: tuple[FinanceRecord, ...]


class TabularImporter(Protocol):
    def load(self, content: bytes, *, column_mapping: Mapping[str, str] | None = None) -> tuple[FinanceRecord, ...]: ...


class _CanonicalRowParser:
    REQUIRED_COLUMNS = {"company", "account", "period", "scenario", "value", "currency"}

    def parse(
        self,
        rows: Iterable[Mapping[str, object]],
        *,
        source_columns: set[str],
        column_mapping: Mapping[str, str] | None = None,
        first_data_row: int = 2,
    ) -> tuple[FinanceRecord, ...]:
        mapping = dict(column_mapping or {})
        required_sources = {mapping.get(name, name) for name in self.REQUIRED_COLUMNS}
        missing = sorted(required_sources - source_columns)
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        records: list[FinanceRecord] = []
        for row_number, row in enumerate(rows, start=first_data_row):
            value_raw = row.get(mapping.get("value", "value"))
            try:
                value = Decimal(str(value_raw).strip())
            except (InvalidOperation, AttributeError) as exc:
                raise ValueError(f"Invalid decimal value at row {row_number}") from exc
            dimensions = tuple(
                sorted(
                    (key.removeprefix("dim_"), str(raw).strip())
                    for key, raw in row.items()
                    if key.startswith("dim_") and raw is not None and str(raw).strip()
                )
            )
            records.append(
                FinanceRecord(
                    company=str(row.get(mapping.get("company", "company"), "")).strip(),
                    account=str(row.get(mapping.get("account", "account"), "")).strip(),
                    period=str(row.get(mapping.get("period", "period"), "")).strip(),
                    scenario=str(row.get(mapping.get("scenario", "scenario"), "")).strip(),
                    value=value,
                    currency=str(row.get(mapping.get("currency", "currency"), "")).strip().upper(),
                    dimensions=dimensions,
                    source_row=row_number,
                )
            )
        return tuple(records)


class CanonicalCsvImporter:
    def __init__(self) -> None:
        self._parser = _CanonicalRowParser()

    def load_text(
        self,
        content: str,
        *,
        column_mapping: Mapping[str, str] | None = None,
    ) -> tuple[FinanceRecord, ...]:
        reader = csv.DictReader(io.StringIO(content))
        return self._parser.parse(
            reader,
            source_columns=set(reader.fieldnames or ()),
            column_mapping=column_mapping,
        )

    def load(
        self,
        content: bytes,
        *,
        column_mapping: Mapping[str, str] | None = None,
    ) -> tuple[FinanceRecord, ...]:
        return self.load_text(content.decode("utf-8-sig"), column_mapping=column_mapping)


class CanonicalExcelImporter:
    def __init__(self) -> None:
        self._parser = _CanonicalRowParser()

    def load(
        self,
        content: bytes,
        *,
        column_mapping: Mapping[str, str] | None = None,
        sheet_name: str | None = None,
    ) -> tuple[FinanceRecord, ...]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            header_values = next(values)
        except StopIteration as exc:
            raise ValueError("Excel workbook contains no rows") from exc
        headers = [str(value).strip() if value is not None else "" for value in header_values]
        if any(not header for header in headers):
            raise ValueError("Excel header contains empty column names")
        rows = (dict(zip(headers, row, strict=True)) for row in values)
        return self._parser.parse(
            rows,
            source_columns=set(headers),
            column_mapping=column_mapping,
        )


class FinanceDataQualityService:
    def validate(
        self,
        records: Iterable[FinanceRecord],
        *,
        allowed_currencies: set[str] | None = None,
        required_dimensions: set[str] | None = None,
    ) -> DataQualityReport:
        materialized = tuple(records)
        findings: list[DataQualityFinding] = []
        seen: set[tuple[object, ...]] = set()
        if not materialized:
            findings.append(
                DataQualityFinding(
                    code="empty_dataset",
                    severity=FindingSeverity.ERROR,
                    message="dataset must contain at least one finance record",
                )
            )
        for record in materialized:
            required = {
                "company": record.company,
                "account": record.account,
                "period": record.period,
                "scenario": record.scenario,
                "currency": record.currency,
            }
            for field, value in required.items():
                if not value:
                    findings.append(
                        DataQualityFinding(
                            code=f"missing_{field}",
                            severity=FindingSeverity.ERROR,
                            message=f"{field} must not be empty",
                            row_number=record.source_row,
                        )
                    )
            if not self._valid_period(record.period):
                findings.append(
                    DataQualityFinding(
                        code="invalid_period",
                        severity=FindingSeverity.ERROR,
                        message="period must use YYYY-MM with month between 01 and 12",
                        row_number=record.source_row,
                    )
                )
            if len(record.currency) != 3 or not record.currency.isalpha():
                findings.append(
                    DataQualityFinding(
                        code="invalid_currency",
                        severity=FindingSeverity.ERROR,
                        message="currency must be a three-letter code",
                        row_number=record.source_row,
                    )
                )
            elif allowed_currencies is not None and record.currency not in allowed_currencies:
                findings.append(
                    DataQualityFinding(
                        code="unsupported_currency",
                        severity=FindingSeverity.ERROR,
                        message=f"currency {record.currency} is not allowed",
                        row_number=record.source_row,
                    )
                )
            dimension_names = {name for name, _ in record.dimensions}
            for dimension in sorted((required_dimensions or set()) - dimension_names):
                findings.append(
                    DataQualityFinding(
                        code="missing_dimension",
                        severity=FindingSeverity.ERROR,
                        message=f"required dimension {dimension} is missing",
                        row_number=record.source_row,
                    )
                )
            key = record.canonical_key()
            if key in seen:
                findings.append(
                    DataQualityFinding(
                        code="duplicate_record",
                        severity=FindingSeverity.ERROR,
                        message="duplicate canonical finance record",
                        row_number=record.source_row,
                    )
                )
            seen.add(key)
        return DataQualityReport(row_count=len(materialized), findings=tuple(findings))

    @staticmethod
    def _valid_period(value: str) -> bool:
        if len(value) != 7 or value[4] != "-":
            return False
        try:
            year = int(value[:4])
            month = int(value[5:])
        except ValueError:
            return False
        return 1900 <= year <= 9999 and 1 <= month <= 12


class DataSnapshotFactory:
    def create(self, records: Iterable[FinanceRecord]) -> DataSnapshot:
        ordered = tuple(sorted(records, key=lambda item: item.canonical_key()))
        payload = [
            {
                **asdict(record),
                "value": format(record.value, "f"),
                "dimensions": list(record.dimensions),
                "source_row": None,
            }
            for record in ordered
        ]
        canonical = json.dumps(payload, sort_keys=True, separators=(",", ":"))
        digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
        return DataSnapshot(
            snapshot_id=f"sha256:{digest}",
            content_hash=digest,
            row_count=len(ordered),
            records=ordered,
        )
